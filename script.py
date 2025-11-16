import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter

def ajustare_coloane(nume_fisier, nume_foaie):
    try:
        wb = openpyxl.load_workbook(nume_fisier) #wb - workbook
        ws = wb[nume_foaie] #ws - worksheet
    except FileNotFoundError:
        print(f"Fisierul {nume_fisier} nu a fost gasit!")
        return
    except KeyError:
        print(f"Foaia de lucru {nume_foaie} nu a fost gasita in cadrul fisierului {nume_fisier}!")
        return

    latimi_maxime = {} #creez un dictionar pentru a stoca latimile maxime
    for row in ws.rows:
        for cell in row:
            index_coloana = cell.column
            if cell.value is not None:
                lungime_curenta = len(str(cell.value))+2
                latimi_maxime[index_coloana] = max(latimi_maxime.get(index_coloana, 0), lungime_curenta)

    for index_coloana, latime in latimi_maxime.items():
        litera_coloana = get_column_letter(index_coloana)
        ws.column_dimensions[litera_coloana].width = latime

    try:
        wb.save(nume_fisier)
        print(f"Fisierul '{nume_fisier}' a fost salvat cu latimile ajustate.")
    except Exception as e:
        print(f"Eroare la salvarea fisierului: {e}")

def main():
    nume_fisier = 'Rezultate - Excel.xlsx'
    nume_worksheet = 'Grupare - Medii de vanzare'
    nume_worksheet2 = 'Pagina suplimentara'
    workbook = xlsxwriter.Workbook('Rezultate - Excel.xlsx', {'nan_inf_to_errors': True})
    worksheet = workbook.add_worksheet('Grupare - Medii de vanzare')
    worksheet2 = workbook.add_worksheet('Pagina suplimentara')
    headers = [
        'TIP AUTOVEHICUL','MARCA', 'MODEL', 'VERSIUNE', 'AN FABRICATIE', 'CAPACITATE CILINDRICA', 'PUTERE','INTERVAL_KM','MEDIE DE VANZARE',
        'COUNT'
    ]
    worksheet.write_row('A1', headers)
    worksheet2.write_row('A1', headers)
    dataset = pd.read_csv('raportVanzari.csv', encoding='utf-8', encoding_errors='ignore')
    if dataset.empty:
        print("Excelul de date este gol, eroare la incarcare.")
        return
    data_check = dataset.head()
    print("\nVerificarea incarcarii datelor...", data_check)

    # Curatare de date, coloana cu vanzari
    dataset['MODEL'] = dataset['MODEL'].str.replace(' ', '', regex=False)
    dataset['VANZARE'] = dataset['VANZARE'].str.replace(',', '', regex=False).str.replace('?', '',regex=False).str.replace('lei', '', regex=False).str.strip().astype(float)
    dataset['CULOARE'] = dataset['CULOARE'].astype(str).str.lower().str.strip()


    median = dataset['VANZARE'].median()
    dataset = dataset.fillna(median)
    nr_valori_lipsa = dataset['VANZARE'].isnull().sum()
    if nr_valori_lipsa > 0:
        print("\nEroare la completare cu mediana.")
        return

    dataset_ordonat= dataset.sort_values(by="AN FABRICATIE", ascending=True)

    #grupare_marca = {marca: dataset[dataset['MARCA'] == marca] for marca in dataset['MARCA'].unique()}
    current_row = 1
    previous_marca = None
    max_km = dataset_ordonat['KM'].max()
    bins = np.arange(0, max_km+30001, 30000)
    dataset_ordonat['KM'] = pd.cut(dataset_ordonat['KM'], bins = bins, right = True, include_lowest = True).astype(str)
    grupare_tip = {tip: dataset_ordonat[dataset_ordonat['TIP AUTOVEHICUL'] == tip] for tip in
                   dataset_ordonat['TIP AUTOVEHICUL'].unique()}

    for tip, df_tip in grupare_tip.items():
        rezultate_combinate = df_tip.groupby([
        'MARCA','MODEL', 'VERSIUNE', 'AN FABRICATIE', 'CAPACITATE CILINDRICA', 'PUTERE', 'KM'
        ])['VANZARE'].agg(['mean', 'count']).reset_index()
        worksheet.write(current_row, 0, tip)

        for index, row in rezultate_combinate.iterrows():
                current_marca = row['MARCA']
                if current_marca != previous_marca:
                    worksheet.write(current_row, 1, row['MARCA'])
                    previous_marca = current_marca
                worksheet.write(current_row, 2, row['MODEL'])
                worksheet.write(current_row, 3, row['VERSIUNE'])
                worksheet.write(current_row, 4, row['AN FABRICATIE'])
                worksheet.write(current_row, 5, row['CAPACITATE CILINDRICA'])
                worksheet.write(current_row, 6, row['PUTERE'])
                worksheet.write(current_row, 7, row['KM'])
                worksheet.write(current_row, 8, row['mean'])
                worksheet.write(current_row, 9, row['count'])

                worksheet2.write(current_row, 0, tip)
                worksheet2.write(current_row, 1, row['MARCA'])
                worksheet2.write(current_row, 2, row['MODEL'])
                worksheet2.write(current_row, 3, row['VERSIUNE'])
                worksheet2.write(current_row, 4, row['AN FABRICATIE'])
                worksheet2.write(current_row, 5, row['CAPACITATE CILINDRICA'])
                worksheet2.write(current_row, 6, row['PUTERE'])
                worksheet.write(current_row, 7, row['KM'])
                worksheet2.write(current_row, 8, row['mean'])
                worksheet2.write(current_row, 9, row['count'])
                current_row += 1



    worksheet.set_column("A:XFD", 25)
    #worksheet.set_column("D1:D230", 25)
    worksheet2.set_column("A:XFD", 25)
    #worksheet2.set_column("D1:D230", 25)
    workbook.close()



if (__name__ == '__main__'):
    main()
